import asyncio
import time
import json
import httpx
import os
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, F, BaseMiddleware
from aiogram.types import Message, CallbackQuery, TelegramObject, InputMediaPhoto, InputMediaVideo, InputMediaAnimation
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.base import BaseStorage, StorageKey
from aiogram.utils.keyboard import InlineKeyboardBuilder
from supabase import create_client
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SECRET")
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))

class SupabaseFSMStorage(BaseStorage):
    """FSM storage backed by Supabase — переживает редеплои Railway."""

    def __init__(self, sb):
        self._db = sb

    def _key(self, key: StorageKey) -> str:
        return f"{key.bot_id}:{key.chat_id}:{key.user_id}:{key.destiny}"

    async def _upsert(self, key: StorageKey, fields: dict) -> None:
        k = self._key(key)
        await asyncio.to_thread(
            lambda: self._db.table("fsm_storage")
                .upsert({"key": k, **fields}, on_conflict="key")
                .execute()
        )

    async def _fetch(self, key: StorageKey) -> dict:
        k = self._key(key)
        res = await asyncio.to_thread(
            lambda: self._db.table("fsm_storage")
                .select("state,data")
                .eq("key", k)
                .execute()
        )
        return res.data[0] if res.data else {}

    async def set_state(self, key: StorageKey, state=None) -> None:
        state_str = state.state if hasattr(state, "state") else state
        await self._upsert(key, {"state": state_str})

    async def get_state(self, key: StorageKey):
        return (await self._fetch(key)).get("state")

    async def set_data(self, key: StorageKey, data: dict) -> None:
        await self._upsert(key, {"data": json.dumps(data, ensure_ascii=False)})

    async def get_data(self, key: StorageKey) -> dict:
        raw = (await self._fetch(key)).get("data")
        if not raw:
            return {}
        try:
            return json.loads(raw)
        except Exception:
            return {}

    async def close(self) -> None:
        pass


bot = Bot(token=BOT_TOKEN)
_user_media_locks: dict[int, asyncio.Lock] = {}
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
storage = SupabaseFSMStorage(supabase)
dp = Dispatcher(storage=storage)


# ============================================================
# Middleware: проверка доступа
# ============================================================

class AccessMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        user = data.get("event_from_user")
        if user:
            res = supabase.table("users").select("id").eq("telegram_id", user.id).execute()
            if not res.data:
                if isinstance(event, Message):
                    await event.answer("⛔️ У вас нет доступа. Обратитесь к администратору.")
                elif isinstance(event, CallbackQuery):
                    await event.answer("⛔️ У вас нет доступа. Обратитесь к администратору.", show_alert=True)
                return
        return await handler(event, data)

dp.message.middleware(AccessMiddleware())
dp.callback_query.middleware(AccessMiddleware())


# ============================================================
# FSM состояния
# ============================================================

class BuyStone(StatesGroup):
    stone_type = State()
    stone_type_custom = State()
    stone_search = State()
    origin = State()
    shape = State()
    shape_custom = State()
    carat = State()
    carat_custom = State()
    color = State()
    clarity = State()
    price = State()
    price_confirm = State()
    currency = State()
    exchange_rate = State()
    supplier = State()
    new_supplier = State()
    confirm = State()
    duplicate_confirm = State()
    photo = State()
    certificate = State()

class SellStone(StatesGroup):
    select_stone = State()
    price = State()
    currency = State()
    exchange_rate = State()
    client = State()
    new_client = State()
    confirm = State()

class TransferStone(StatesGroup):
    select_stone = State()
    partner = State()
    new_partner = State()
    confirm = State()

class ReturnStone(StatesGroup):
    select_stone  = State()
    counterparty  = State()
    reason        = State()
    confirm       = State()


# ============================================================
# PIN — сессии и блокировка
# ============================================================

PIN_SESSION_SECONDS = 900
PIN_MAX_ATTEMPTS    = 3
PIN_LOCKOUT_SECONDS = 600

_pin_sessions: dict[int, float] = {}
_pin_failed:   dict[int, int]   = {}
_pin_locked:   dict[int, float] = {}


class PinVerify(StatesGroup):
    enter = State()

class PinSetup(StatesGroup):
    enter_new = State()
    confirm   = State()

class PinChange(StatesGroup):
    verify_old  = State()
    enter_new   = State()
    confirm_new = State()


def _pin_session_active(user_id: int) -> bool:
    return time.time() < _pin_sessions.get(user_id, 0)

def _start_pin_session(user_id: int) -> None:
    _pin_sessions[user_id] = time.time() + PIN_SESSION_SECONDS
    _pin_failed.pop(user_id, None)
    _pin_locked.pop(user_id, None)

def _pin_lockout_remaining(user_id: int) -> float:
    return max(0.0, _pin_locked.get(user_id, 0) - time.time())

def _record_failed_pin(user_id: int) -> int:
    count = _pin_failed.get(user_id, 0) + 1
    _pin_failed[user_id] = count
    if count >= PIN_MAX_ATTEMPTS:
        _pin_locked[user_id] = time.time() + PIN_LOCKOUT_SECONDS
    return count


_PIN_BYPASS_STATES = frozenset({
    PinVerify.enter.state,
    PinSetup.enter_new.state, PinSetup.confirm.state,
    PinChange.verify_old.state, PinChange.enter_new.state, PinChange.confirm_new.state,
})


class PinMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        user = data.get("event_from_user")
        if not user:
            return await handler(event, data)

        state: FSMContext = data.get("state")
        if state and (await state.get_state()) in _PIN_BYPASS_STATES:
            return await handler(event, data)

        if _pin_session_active(user.id):
            return await handler(event, data)

        remaining = _pin_lockout_remaining(user.id)
        if remaining > 0:
            mins = int(remaining // 60) + 1
            text = f"🔒 Слишком много неверных попыток. Попробуй через {mins} мин."
            if isinstance(event, Message):
                await event.answer(text)
            elif isinstance(event, CallbackQuery):
                await event.answer(text, show_alert=True)
            return

        res = supabase.table("users").select("pin_code").eq("telegram_id", user.id).execute()
        if not res.data:
            return

        if state:
            await state.clear()

        pin_code = (res.data[0].get("pin_code") or "").strip()
        if not pin_code:
            await state.set_state(PinSetup.enter_new)
            text = "🔐 *Добро пожаловать!*\n\nУстанови PIN-код — 4 цифры:"
            pm = "Markdown"
        else:
            await state.set_state(PinVerify.enter)
            text = "🔐 Введи PIN-код:"
            pm = None

        if isinstance(event, Message):
            await event.answer(text, parse_mode=pm)
        elif isinstance(event, CallbackQuery):
            await event.answer()
            await event.message.answer(text, parse_mode=pm)


dp.message.middleware(PinMiddleware())
dp.callback_query.middleware(PinMiddleware())


# ============================================================
# Вспомогательные функции
# ============================================================

def get_status_emoji(status):
    return {"in_stock": "🟢", "at_partner": "🔵", "reserved": "🟡",
            "sent_to_client": "🟠", "in_jewelry": "💍", "sold": "✅",
            "written_off": "❌"}.get(status, "⚪")

def get_status_name(status):
    return {"in_stock": "В наличии", "at_partner": "У партнёра",
            "reserved": "В резерве", "sent_to_client": "У клиента",
            "in_jewelry": "В изделии", "sold": "Продан",
            "written_off": "Списан"}.get(status, status)

def get_user_id(telegram_id):
    res = supabase.table("users").select("id").eq("telegram_id", telegram_id).execute()
    return res.data[0]["id"] if res.data else None

def abbr_code(code: str) -> str:
    parts = code.split("-")
    if len(parts) == 3 and len(parts[1]) == 4:
        parts[1] = parts[1][2:]
    return "-".join(parts)

def abbr_color(c: str) -> str:
    return c if len(c) <= 5 else c[:4] + "."

_SEARCH_STONES = [
    "Аквамарин", "Александрит", "Аметист", "Гранат", "Жемчуг",
    "Опал", "Сапфир", "Танзанит", "Топаз", "Турмалин",
    "Хризолит", "Цаворит", "Янтарь",
]


def abbr_type(full_type: str) -> str:
    type_map = {"diamond": "Бри", "emerald": "Изу", "ruby": "Руб", "spinel": "Шпи"}
    origin_map = {"Природный": "прир", "Синтетический": "синт"}
    if "(" in full_type:
        name, rest = full_type.split("(", 1)
        origin = rest.strip(" )")
        abbr = type_map.get(name.strip(), name.strip()[:3])
        return f"{abbr}/{origin_map.get(origin, origin[:4])}"
    return type_map.get(full_type.strip(), full_type.strip()[:3])

def fmt_stone_btn(s: dict) -> str:
    return f"{abbr_type(s['stone_type'])} {s['carat']}кар · {abbr_code(s['stone_code'])}"

def next_stone_code():
    year = date.today().year
    prefix = f"ST-{year}-"
    res = supabase.table("stones").select("stone_code").like("stone_code", f"{prefix}%").execute()
    if not res.data:
        return f"{prefix}001"
    max_num = 0
    for row in res.data:
        try:
            num = int(row["stone_code"].replace(prefix, ""))
            if num > max_num:
                max_num = num
        except:
            pass
    return f"{prefix}{max_num + 1:03d}"


# ============================================================
# Главное меню
# ============================================================

def main_keyboard():
    kb = InlineKeyboardBuilder()
    kb.button(text="💎 Купили", callback_data="action_buy")
    kb.button(text="💰 Продали", callback_data="action_sell")
    kb.button(text="📤 Ювелиру", callback_data="action_transfer")
    kb.button(text="↩️ Возврат", callback_data="action_return")
    kb.button(text="📋 Склад", callback_data="action_inventory")
    kb.button(text="💵 Итого", callback_data="action_total")
    kb.button(text="📊 Экспорт", callback_data="action_export")
    kb.button(text="👁 Посмотреть", callback_data="action_view")
    kb.adjust(2, 2, 2, 2)
    return kb.as_markup()


# ============================================================
# AI — Claude
# ============================================================

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
ANTHROPIC_MODEL   = "claude-sonnet-4-6"

_AI_TOOLS = [
    {
        "name": "get_inventory",
        "description": "Получить все камни в наличии (не проданные и не списанные) со статусами и стоимостью. "
                       "Поддерживает текстовый поиск по типу камня через search_query.",
        "input_schema": {
            "type": "object",
            "properties": {
                "search_query": {"type": "string", "description": "Текстовый поиск по stone_type (ILIKE). Передавай и русское, и английское слово через запятую, например: сапфир,sapphire"},
            },
            "required": [],
        },
    },
    {
        "name": "get_stone",
        "description": "Найти камень(ни) по коду или характеристикам.",
        "input_schema": {
            "type": "object",
            "properties": {
                "stone_code": {"type": "string",  "description": "Код камня, например ST-2026-001"},
                "stone_type": {"type": "string",  "description": "Тип: diamond, emerald, ruby, spinel или часть названия"},
                "min_carat":  {"type": "number",  "description": "Минимальный вес в каратах"},
                "max_carat":  {"type": "number",  "description": "Максимальный вес в каратах"},
                "status":     {"type": "string",  "description": "in_stock / at_partner / reserved / sent_to_client / in_jewelry / sold / written_off"},
                "color":      {"type": "string",  "description": "Цвет камня"},
                "clarity":    {"type": "string",  "description": "Чистота камня"},
            },
            "required": [],
        },
    },
    {
        "name": "get_operations",
        "description": "Получить последние операции: покупки, продажи, передачи ювелиру.",
        "input_schema": {
            "type": "object",
            "properties": {
                "operation_type": {"type": "string",  "description": "purchase_stone / sale_stone / transfer_to_partner"},
                "limit":          {"type": "integer", "description": "Кол-во записей, по умолчанию 20"},
                "date_from":      {"type": "string",  "description": "С даты YYYY-MM-DD"},
                "date_to":        {"type": "string",  "description": "По дату YYYY-MM-DD"},
            },
            "required": [],
        },
    },
    {
        "name": "get_total",
        "description": "Итоговая стоимость склада: количество камней, суммарный вес, общая стоимость в USD.",
        "input_schema": {"type": "object", "properties": {}, "required": []},
    },
]


def _ai_get_inventory(search_query: str = None) -> str:
    q = supabase.table("v_stone_current_value") \
        .select("stone_code,stone_type,carat,color,clarity,status,current_value_usd") \
        .not_.in_("status", ["sold", "written_off"])
    if search_query:
        terms = [t.strip() for t in search_query.split(",") if t.strip()]
        if terms:
            or_filter = ",".join(f"stone_type.ilike.%{t}%" for t in terms)
            q = q.or_(or_filter)
    return json.dumps(q.order("carat", desc=True).execute().data or [], ensure_ascii=False)


def _ai_get_stone(stone_code=None, stone_type=None, min_carat=None, max_carat=None,
                  status=None, color=None, clarity=None) -> str:
    q = supabase.table("v_stone_current_value").select(
        "stone_code,stone_type,shape,carat,color,clarity,status,"
        "current_value_usd,purchase_price,purchase_currency,purchase_date"
    )
    if stone_code:
        q = q.ilike("stone_code", f"%{stone_code}%")
    if stone_type:
        q = q.ilike("stone_type", f"%{stone_type}%")
    if min_carat is not None:
        q = q.gte("carat", min_carat)
    if max_carat is not None:
        q = q.lte("carat", max_carat)
    if status:
        q = q.eq("status", status)
    if color:
        q = q.ilike("color", f"%{color}%")
    if clarity:
        q = q.ilike("clarity", f"%{clarity}%")
    return json.dumps(q.order("carat", desc=True).execute().data or [], ensure_ascii=False)


def _ai_get_operations(operation_type=None, limit=20, date_from=None, date_to=None) -> str:
    q = supabase.table("operations").select(
        "operation_type,amount,currency,amount_usd,created_at,entity_id,counterparty_id,notes"
    ).order("created_at", desc=True).limit(int(limit) if limit else 20)
    if operation_type:
        q = q.eq("operation_type", operation_type)
    if date_from:
        q = q.gte("created_at", date_from)
    if date_to:
        q = q.lte("created_at", date_to + "T23:59:59")
    ops = q.execute().data or []

    stone_ids = list({op["entity_id"] for op in ops if op.get("entity_id")})
    cp_ids    = list({op["counterparty_id"] for op in ops if op.get("counterparty_id")})
    stones_map: dict = {}
    if stone_ids:
        s = supabase.table("stones").select("id,stone_code,stone_type,carat").in_("id", stone_ids).execute()
        stones_map = {r["id"]: r for r in (s.data or [])}
    cp_map: dict = {}
    if cp_ids:
        c = supabase.table("counterparties").select("id,name").in_("id", cp_ids).execute()
        cp_map = {r["id"]: r["name"] for r in (c.data or [])}

    for op in ops:
        op["stone"] = stones_map.get(op.get("entity_id"))
        op["counterparty_name"] = cp_map.get(op.get("counterparty_id"))
    return json.dumps(ops, ensure_ascii=False, default=str)


def _ai_get_total() -> str:
    rows = supabase.table("v_stone_current_value") \
        .select("current_value_usd,carat") \
        .not_.in_("status", ["sold", "written_off"]).execute().data or []
    return json.dumps({
        "count":       len(rows),
        "total_carat": round(sum(r["carat"] or 0 for r in rows), 2),
        "total_usd":   round(sum(r["current_value_usd"] or 0 for r in rows), 2),
    }, ensure_ascii=False)


def _ai_run_tool(name: str, tool_input: dict) -> str:
    if name == "get_inventory":
        return _ai_get_inventory(**tool_input)
    if name == "get_stone":
        return _ai_get_stone(**tool_input)
    if name == "get_operations":
        return _ai_get_operations(**tool_input)
    if name == "get_total":
        return _ai_get_total()
    return json.dumps({"error": f"unknown tool: {name}"})


async def ask_claude(user_text: str) -> str:
    system_prompt = (
        "Ты помощник для учёта драгоценных камней. У тебя есть доступ к данным склада через инструменты. "
        "Отвечай на русском языке, кратко и по делу. "
        "Для получения данных всегда используй инструменты — никогда не выдумывай информацию. "
        "ЗАПРЕЩЕНО использовать таблицы Markdown (| col | col |) — Telegram их не поддерживает. "
        "Список камней выводи отдельными строками в формате: "
        "💎 ST-2026-001 — Рубин, 2.10 кар, Red, VS2 — $6,207. "
        "Статусы камней всегда переводи на русский: "
        "in_stock=в наличии, at_partner=у партнёра, reserved=в резерве, "
        "sold=продан, written_off=списан. "
        "При поиске по типу камня всегда используй ILIKE (частичное совпадение): "
        "передавай в search_query и русское, и английское название через запятую. "
        "Примеры: сапфир→'сапфир,sapphire'; синтетический бриллиант→'синтет,synthetic,diamond'; "
        "рубин→'рубин,ruby'; изумруд→'изумруд,emerald'; шпинель→'шпинель,spinel'. "
        "Типы операций: purchase_stone=покупка, sale_stone=продажа, "
        "transfer_to_partner=передача ювелиру, "
        "return_from_client=возврат от клиента, return_from_partner=возврат от партнёра. "
        f"Сегодня {date.today().strftime('%d.%m.%Y')}."
    )
    messages = [{"role": "user", "content": user_text}]

    async with httpx.AsyncClient(timeout=30) as client:
        for _ in range(5):
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": ANTHROPIC_MODEL,
                    "max_tokens": 1024,
                    "system": [{"type": "text", "text": system_prompt,
                                "cache_control": {"type": "ephemeral"}}],
                    "tools": _AI_TOOLS,
                    "messages": messages,
                },
            )
            resp.raise_for_status()
            body = resp.json()

            if body["stop_reason"] == "end_turn":
                for block in body["content"]:
                    if block["type"] == "text":
                        return block["text"]
                return "Нет ответа."

            if body["stop_reason"] == "tool_use":
                messages.append({"role": "assistant", "content": body["content"]})
                tool_results = []
                for block in body["content"]:
                    if block["type"] == "tool_use":
                        result = await asyncio.to_thread(
                            _ai_run_tool, block["name"], block.get("input", {})
                        )
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block["id"],
                            "content": result,
                        })
                messages.append({"role": "user", "content": tool_results})
                continue
            break

    return "Не удалось получить ответ."


@dp.message(Command("history"))
async def cmd_history(message: Message, state: FSMContext):
    await state.clear()
    text = _build_history_text()
    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await message.answer(
        text or "История операций пуста.",
        reply_markup=kb.as_markup(), parse_mode="Markdown" if text else None)


@dp.message(Command("adduser"))
async def cmd_adduser(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Нет доступа.")
        return

    args = message.text.split(maxsplit=2)
    if len(args) < 3:
        await message.answer(
            "Использование: /adduser <telegram_id> <имя> [роль]\n\n"
            "Роли: owner, partner\n"
            "Пример: /adduser 123456789 Иван manager")
        return

    try:
        tg_id = int(args[1])
    except ValueError:
        await message.answer("❌ telegram_id должен быть числом.")
        return

    name = args[2]
    role_parts = name.rsplit(maxsplit=1)
    valid_roles = {"owner", "partner"}
    if len(role_parts) == 2 and role_parts[1] in valid_roles:
        name = role_parts[0]
        role = role_parts[1]
    else:
        role = "partner"

    existing = supabase.table("users").select("id").eq("telegram_id", tg_id).execute()
    if existing.data:
        await message.answer(f"⚠️ Пользователь с ID {tg_id} уже есть в базе.")
        return

    try:
        supabase.table("users").insert({
            "telegram_id": tg_id,
            "name": name,
            "role": role,
            "is_active": True,
        }).execute()
        await message.answer(f"✅ Пользователь добавлен:\nID: {tg_id}\nИмя: {name}\nРоль: {role}")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")


@dp.message(Command("addpartner"))
async def cmd_addpartner(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Нет доступа.")
        return

    args = message.text.split(maxsplit=2)
    valid_types = {"partner", "supplier", "client"}

    if len(args) < 3 or args[1] not in valid_types:
        await message.answer(
            "Использование: /addpartner <тип> <имя>\n\n"
            "Типы: partner, supplier, client\n\n"
            "Примеры:\n"
            "/addpartner partner Мастер Ли\n"
            "/addpartner supplier Поставщик Чэн\n"
            "/addpartner client Клиент Иванов")
        return

    cp_type = args[1]
    name = args[2].strip()

    existing = supabase.table("counterparties").select("id").eq("name", name).eq("type", cp_type).execute()
    if existing.data:
        await message.answer(f"⚠️ {cp_type} с именем «{name}» уже есть в базе.")
        return

    type_emoji = {"partner": "📤", "supplier": "🏪", "client": "👤"}
    try:
        supabase.table("counterparties").insert({
            "name": name,
            "type": cp_type,
            "is_active": True,
        }).execute()
        await message.answer(
            f"✅ Добавлен {cp_type}:\n{type_emoji[cp_type]} {name}")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")


@dp.message(Command("listusers"))
async def cmd_listusers(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Нет доступа.")
        return

    res = supabase.table("users").select("telegram_id,name,role,is_active").order("created_at").execute()
    if not res.data:
        await message.answer("Список пользователей пуст.")
        return

    role_emoji = {"owner": "👑", "partner": "🤝"}
    lines = ["👥 *Пользователи:*\n"]
    for u in res.data:
        emoji = role_emoji.get(u["role"], "👤")
        active = "✅" if u["is_active"] else "🚫"
        lines.append(f"{active} {emoji} *{u['name']}*\n   ID: `{u['telegram_id']}` · {u['role']}")

    await message.answer("\n\n".join(lines), parse_mode="Markdown")


@dp.message(Command("deluser"))
async def cmd_deluser(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Нет доступа.")
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer(
            "Использование: /deluser <telegram_id>\n\n"
            "Пример: /deluser 123456789")
        return

    try:
        tg_id = int(args[1])
    except ValueError:
        await message.answer("❌ telegram_id должен быть числом.")
        return

    if tg_id == ADMIN_ID:
        await message.answer("❌ Нельзя удалить самого себя.")
        return

    existing = supabase.table("users").select("id,name,role").eq("telegram_id", tg_id).execute()
    if not existing.data:
        await message.answer(f"⚠️ Пользователь с ID {tg_id} не найден.")
        return

    user = existing.data[0]
    try:
        supabase.table("users").delete().eq("telegram_id", tg_id).execute()
        await message.answer(
            f"✅ Пользователь удалён:\nID: {tg_id}\nИмя: {user['name']}\nРоль: {user['role']}")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")


@dp.message(Command("setcommands"))
async def cmd_setcommands(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Нет доступа.")
        return
    from aiogram.types import BotCommand
    commands = [
        BotCommand(command="start",       description="Главное меню"),
        BotCommand(command="menu",        description="Главное меню"),
        BotCommand(command="stone",       description="Карточка камня: /stone ST-2026-001"),
        BotCommand(command="history",     description="Последние 10 операций"),
        BotCommand(command="cancel",      description="Отменить текущее действие"),
        BotCommand(command="adduser",     description="Добавить пользователя (admin)"),
        BotCommand(command="deluser",     description="Удалить пользователя (admin)"),
        BotCommand(command="listusers",   description="Список пользователей (admin)"),
        BotCommand(command="setpin",       description="Сменить PIN-код"),
        BotCommand(command="addpartner",  description="Добавить контрагента (admin)"),
        BotCommand(command="setcommands", description="Обновить меню команд (admin)"),
    ]
    await bot.set_my_commands(commands)
    await message.answer("✅ Команды зарегистрированы в меню Telegram.")


# ============================================================
# PIN — обработчики
# ============================================================

@dp.message(PinSetup.enter_new)
async def pin_setup_new(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    pin = message.text.strip() if message.text else ""
    if not pin.isdigit() or len(pin) != 4:
        await message.answer("❌ PIN должен быть ровно 4 цифры. Попробуй ещё раз:")
        return
    await state.update_data(new_pin=pin)
    await state.set_state(PinSetup.confirm)
    await message.answer("🔐 Повтори PIN:")


@dp.message(PinSetup.confirm)
async def pin_setup_confirm(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    data = await state.get_data()
    pin = message.text.strip() if message.text else ""
    if pin != data.get("new_pin"):
        await state.set_state(PinSetup.enter_new)
        await message.answer("❌ PIN не совпадает. Введи заново (4 цифры):")
        return
    try:
        supabase.table("users").update({"pin_code": pin}) \
            .eq("telegram_id", message.from_user.id).execute()
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        return
    _start_pin_session(message.from_user.id)
    await state.clear()
    await message.answer("✅ PIN установлен!\n\n💎 *Jewelry AI*\n\nВыбери действие:",
                         reply_markup=main_keyboard(), parse_mode="Markdown")


@dp.message(PinVerify.enter)
async def pin_verify(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    user_id = message.from_user.id
    remaining = _pin_lockout_remaining(user_id)
    if remaining > 0:
        mins = int(remaining // 60) + 1
        await message.answer(f"🔒 Аккаунт заблокирован. Попробуй через {mins} мин.")
        return
    pin = message.text.strip() if message.text else ""
    res = supabase.table("users").select("pin_code").eq("telegram_id", user_id).execute()
    correct = res.data[0]["pin_code"] if res.data else None
    if pin == correct:
        _start_pin_session(user_id)
        await state.clear()
        await message.answer("💎 *Jewelry AI*\n\nВыбери действие:",
                             reply_markup=main_keyboard(), parse_mode="Markdown")
    else:
        count = _record_failed_pin(user_id)
        if _pin_lockout_remaining(user_id) > 0:
            await message.answer(
                f"❌ Неверный PIN. Аккаунт заблокирован на {PIN_LOCKOUT_SECONDS // 60} мин.")
        else:
            left = PIN_MAX_ATTEMPTS - count
            await message.answer(f"❌ Неверный PIN. Осталось попыток: {left}")


@dp.message(Command("setpin"))
async def cmd_setpin(message: Message, state: FSMContext):
    res = supabase.table("users").select("pin_code").eq("telegram_id", message.from_user.id).execute()
    has_pin = bool(res.data and (res.data[0].get("pin_code") or "").strip())
    if has_pin:
        await state.set_state(PinChange.verify_old)
        await message.answer("🔐 Введи текущий PIN:")
    else:
        await state.set_state(PinSetup.enter_new)
        await message.answer("🔐 Установи новый PIN (4 цифры):")


@dp.message(PinChange.verify_old)
async def pin_change_old(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    user_id = message.from_user.id
    remaining = _pin_lockout_remaining(user_id)
    if remaining > 0:
        mins = int(remaining // 60) + 1
        await message.answer(f"🔒 Аккаунт заблокирован. Попробуй через {mins} мин.")
        await state.clear()
        return
    pin = message.text.strip() if message.text else ""
    res = supabase.table("users").select("pin_code").eq("telegram_id", user_id).execute()
    correct = res.data[0]["pin_code"] if res.data else None
    if pin == correct:
        _pin_failed.pop(user_id, None)
        _pin_locked.pop(user_id, None)
        await state.set_state(PinChange.enter_new)
        await message.answer("🔐 Введи новый PIN (4 цифры):")
    else:
        count = _record_failed_pin(user_id)
        if _pin_lockout_remaining(user_id) > 0:
            await message.answer(
                f"❌ Неверный PIN. Аккаунт заблокирован на {PIN_LOCKOUT_SECONDS // 60} мин.")
            await state.clear()
        else:
            left = PIN_MAX_ATTEMPTS - count
            await message.answer(f"❌ Неверный PIN. Осталось попыток: {left}")


@dp.message(PinChange.enter_new)
async def pin_change_new(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    pin = message.text.strip() if message.text else ""
    if not pin.isdigit() or len(pin) != 4:
        await message.answer("❌ PIN должен быть 4 цифры:")
        return
    await state.update_data(new_pin=pin)
    await state.set_state(PinChange.confirm_new)
    await message.answer("🔐 Повтори новый PIN:")


@dp.message(PinChange.confirm_new)
async def pin_change_confirm(message: Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    data = await state.get_data()
    pin = message.text.strip() if message.text else ""
    if pin != data.get("new_pin"):
        await state.set_state(PinChange.enter_new)
        await message.answer("❌ PIN не совпадает. Введи новый PIN (4 цифры):")
        return
    try:
        supabase.table("users").update({"pin_code": pin}) \
            .eq("telegram_id", message.from_user.id).execute()
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")
        return
    _start_pin_session(message.from_user.id)
    await state.clear()
    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await message.answer("✅ PIN успешно изменён!", reply_markup=kb.as_markup())


@dp.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    current = await state.get_state()
    if current in {PinVerify.enter.state, PinSetup.enter_new.state, PinSetup.confirm.state}:
        await message.answer("🔐 Введи PIN-код для продолжения:")
        return
    await state.clear()
    await message.answer("❌ Отменено.\n\n💎 *Jewelry AI*\n\nВыбери действие:",
                        reply_markup=main_keyboard(), parse_mode="Markdown")

@dp.message(Command("start"))
@dp.message(Command("menu"))
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("💎 *Jewelry AI*\n\nВыбери действие:",
                        reply_markup=main_keyboard(), parse_mode="Markdown")


# ============================================================
# СКЛАД
# ============================================================

def _build_history_text() -> str | None:
    res = supabase.table("operations") \
        .select("operation_type,amount,currency,amount_usd,created_at,entity_id,counterparty_id,notes") \
        .order("created_at", desc=True).limit(10).execute()
    if not res.data:
        return None

    stone_ids = list({op["entity_id"] for op in res.data if op.get("entity_id")})
    cp_ids    = list({op["counterparty_id"] for op in res.data if op.get("counterparty_id")})

    stones_map = {}
    if stone_ids:
        s_res = supabase.table("stones").select("id,stone_code,stone_type,carat") \
            .in_("id", stone_ids).execute()
        stones_map = {s["id"]: s for s in (s_res.data or [])}

    cp_map = {}
    if cp_ids:
        c_res = supabase.table("counterparties").select("id,name") \
            .in_("id", cp_ids).execute()
        cp_map = {c["id"]: c["name"] for c in (c_res.data or [])}

    op_names = {
        "sale_stone": "💰 Продажа",
        "transfer_to_partner": "📤 Ювелиру",
        "purchase_stone":        "💎 Покупка",
        "return_from_client":    "↩️ Возврат от клиента",
        "return_from_partner":   "↩️ Возврат от партнёра",
    }

    lines = ["📋 *Последние 10 операций:*\n"]
    for op in res.data:
        op_type = op_names.get(op["operation_type"], op["operation_type"])
        stone = stones_map.get(op.get("entity_id")) or {}
        sc = stone.get("stone_code", "")
        stone_info = (f"{abbr_type(stone['stone_type'])} {stone.get('carat','')}кар · {abbr_code(sc)}"
                      if sc else "—")
        cp_name = cp_map.get(op.get("counterparty_id"), "")
        amount_str = f"{op['amount']:,.0f} {op['currency']} ({op['amount_usd']:,.0f} USD)" \
            if op.get("amount") else ""
        date_str = op["created_at"][:10] if op.get("created_at") else ""
        line = f"{op_type} · {date_str}\n  {stone_info}"
        if cp_name:
            line += f" → {cp_name}"
        if amount_str:
            line += f"\n  {amount_str}"
        if op.get("notes"):
            line += f"\n  📝 {op['notes']}"
        lines.append(line)

    return "\n\n".join(lines)


@dp.callback_query(F.data == "action_history")
async def show_history(callback: CallbackQuery):
    text = _build_history_text()
    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await callback.message.edit_text(
        text or "История операций пуста.",
        reply_markup=kb.as_markup(), parse_mode="Markdown" if text else None)


@dp.callback_query(F.data == "action_inventory")
async def show_inventory(callback: CallbackQuery):
    data = supabase.table("v_stone_current_value") \
        .select("stone_code,stone_type,carat,color,clarity,status,current_value_usd") \
        .not_.in_("status", ["sold", "written_off"]) \
        .order("carat", desc=True).execute()

    if not data.data:
        await callback.message.edit_text("Склад пуст.", reply_markup=main_keyboard())
        return

    lines = ["💎 *Склад*\n"]
    for s in data.data:
        emoji = get_status_emoji(s["status"])
        status = get_status_name(s["status"])
        color = abbr_color(s.get("color") or "")
        clarity = s.get("clarity") or ""
        chars = " · ".join(filter(None, [color, clarity]))
        lines.append(
            f"{emoji} *{abbr_type(s['stone_type'])}* {s['carat']}кар"
            + (f" · {chars}" if chars else "")
            + f"\n   {status} · {abbr_code(s['stone_code'])} · {s['current_value_usd']:,.0f} USD"
        )

    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await callback.message.edit_text("\n".join(lines),
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


# ============================================================
# ИТОГО
# ============================================================

@dp.callback_query(F.data == "action_total")
async def show_total(callback: CallbackQuery):
    data = supabase.table("v_stone_current_value") \
        .select("current_value_usd,carat,status") \
        .not_.in_("status", ["sold", "written_off"]).execute()

    total_usd = sum(s["current_value_usd"] or 0 for s in data.data)
    total_carat = sum(s["carat"] or 0 for s in data.data)
    count = len(data.data)

    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await callback.message.edit_text(
        f"📊 *Стоимость склада*\n\nКамней: {count}\nКаратов: {total_carat:.2f}\n"
        f"Итого: *{total_usd:,.0f} USD*",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


# ============================================================
# ЭКСПОРТ
# ============================================================

@dp.callback_query(F.data == "action_export")
async def export_stones(callback: CallbackQuery):
    await callback.answer()
    await callback.message.edit_text("⏳ Формирую файл...")

    stones = supabase.table("v_stone_current_value") \
        .select("stone_code,stone_type,shape,carat,color,clarity,status,"
                "purchase_price,purchase_currency,current_value_usd,purchase_date") \
        .order("purchase_date", desc=True).execute().data or []

    status_map = {"in_stock": "В наличии", "at_partner": "У партнёра",
                  "reserved": "В резерве", "sent_to_client": "У клиента",
                  "in_jewelry": "В изделии", "sold": "Продан", "written_off": "Списан"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Камни"

    headers = ["Код", "Тип", "Форма", "Вес", "Цвет",
               "Чистота", "Статус", "Цена покупки", "Валюта", "Стоимость USD", "Дата покупки"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [14, 20, 14, 8, 8, 10, 14, 14, 8, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for r, s in enumerate(stones, 2):
        ws.append([
            s.get("stone_code", ""),
            s.get("stone_type", ""),
            s.get("shape", ""),
            s.get("carat"),
            s.get("color", ""),
            s.get("clarity", ""),
            status_map.get(s.get("status", ""), s.get("status", "")),
            s.get("purchase_price"),
            s.get("purchase_currency", ""),
            s.get("current_value_usd"),
            s.get("purchase_date", ""),
        ])
        if r % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D6E4F0")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from aiogram.types import BufferedInputFile
    filename = f"stones_{date.today().strftime('%Y%m%d')}.xlsx"
    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"📊 Экспорт: {len(stones)} камней · {date.today()}")
    await callback.message.edit_text("✅ Файл отправлен.", reply_markup=kb.as_markup())


# ============================================================
# КУПИЛИ — шаг 1: тип камня
# ============================================================

@dp.callback_query(F.data == "action_buy")
async def buy_step1_type(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    kb = InlineKeyboardBuilder()
    kb.button(text="💎 Бриллиант", callback_data="type_diamond")
    kb.button(text="💚 Изумруд", callback_data="type_emerald")
    kb.button(text="🔴 Рубин", callback_data="type_ruby")
    kb.button(text="🩷 Шпинель", callback_data="type_spinel")
    kb.button(text="🔍 Найти камень", callback_data="type_search")
    kb.button(text="◀️ Назад", callback_data="back_menu")
    kb.adjust(2, 2, 1, 1)
    await state.set_state(BuyStone.stone_type)
    await callback.message.edit_text("💎 *Тип камня?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


# ============================================================
# КУПИЛИ — шаг 2: происхождение
# ============================================================

@dp.callback_query(BuyStone.stone_type, F.data.in_({"type_diamond", "type_emerald", "type_ruby", "type_spinel"}))
async def buy_step2_origin(callback: CallbackQuery, state: FSMContext):
    type_map = {"type_diamond": "diamond", "type_emerald": "emerald",
                "type_ruby": "ruby", "type_spinel": "spinel"}
    name_map = {"diamond": "Бриллиант", "emerald": "Изумруд",
                "ruby": "Рубин", "spinel": "Шпинель"}
    stone_type = type_map[callback.data]
    await state.update_data(stone_type=stone_type)

    kb = InlineKeyboardBuilder()
    kb.button(text="🌍 Природный", callback_data="origin_natural")
    kb.button(text="🔬 Синтетический", callback_data="origin_synthetic")
    kb.button(text="◀️ Назад", callback_data="action_buy")
    kb.adjust(2, 1)
    await state.set_state(BuyStone.origin)
    await callback.message.edit_text(
        f"*{name_map[stone_type]}*\n\n🌍 Происхождение?",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


# ============================================================
# КУПИЛИ — шаг 3: форма
# ============================================================

@dp.callback_query(BuyStone.stone_type, F.data == "type_custom")
async def buy_step1_type_custom(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.stone_type_custom)
    await callback.message.edit_text("✏️ Введи название камня:")


@dp.message(BuyStone.stone_type_custom)
async def buy_step1_type_custom_text(message: Message, state: FSMContext):
    stone_type = message.text.strip()
    await state.update_data(stone_type=stone_type)
    kb = InlineKeyboardBuilder()
    kb.button(text="🌍 Природный", callback_data="origin_natural")
    kb.button(text="🔬 Синтетический", callback_data="origin_synthetic")
    kb.button(text="◀️ Назад", callback_data="action_buy")
    kb.adjust(2, 1)
    await state.set_state(BuyStone.origin)
    await message.answer(f"*{stone_type}*\n\n🌍 Происхождение?",
                         reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.stone_type, F.data == "type_search")
async def buy_step1_type_search(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.stone_search)
    await callback.message.edit_text("🔍 Напиши название камня или первые буквы:")


@dp.message(BuyStone.stone_search)
async def buy_step1_search_query(message: Message, state: FSMContext):
    query = message.text.strip().lower()
    matches = [s for s in _SEARCH_STONES if s.lower().startswith(query)]
    if not matches:
        matches = [s for s in _SEARCH_STONES if query in s.lower()]

    kb = InlineKeyboardBuilder()
    if matches:
        for stone in matches:
            kb.button(text=stone, callback_data=f"found_{stone}")
        kb.button(text="✏️ Ввести вручную", callback_data="manual_input")
        kb.adjust(2)
        await message.answer("🔍 Выбери камень:", reply_markup=kb.as_markup())
    else:
        kb.button(text="✏️ Ввести вручную", callback_data="manual_input")
        kb.button(text="◀️ Назад", callback_data="action_buy")
        kb.adjust(1)
        await message.answer(
            "❌ Камень не найден. Попробуй другой запрос или введи название вручную:",
            reply_markup=kb.as_markup()
        )


@dp.callback_query(BuyStone.stone_search, F.data.startswith("found_"))
async def buy_step1_found_selected(callback: CallbackQuery, state: FSMContext):
    stone_name = callback.data.replace("found_", "")
    await state.update_data(stone_type=stone_name)
    kb = InlineKeyboardBuilder()
    kb.button(text="🌍 Природный", callback_data="origin_natural")
    kb.button(text="🔬 Синтетический", callback_data="origin_synthetic")
    kb.button(text="◀️ Назад", callback_data="action_buy")
    kb.adjust(2, 1)
    await state.set_state(BuyStone.origin)
    await callback.message.edit_text(
        f"*{stone_name}*\n\n🌍 Происхождение?",
        reply_markup=kb.as_markup(), parse_mode="Markdown"
    )


@dp.callback_query(BuyStone.stone_search, F.data == "manual_input")
async def buy_step1_search_manual(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.stone_type_custom)
    await callback.message.edit_text("✏️ Введи название камня:")


@dp.callback_query(BuyStone.origin, F.data.startswith("origin_"))
async def buy_step3_shape(callback: CallbackQuery, state: FSMContext):
    origin = "Природный" if callback.data == "origin_natural" else "Синтетический"
    await state.update_data(origin=origin)

    kb = InlineKeyboardBuilder()
    for shape in ["Round", "Oval", "Cushion", "Pear", "Princess", "Emerald cut", "Marquise"]:
        kb.button(text=shape, callback_data=f"shape_{shape}")
    kb.button(text="✏️ Другая", callback_data="shape_custom")
    kb.button(text="◀️ Назад", callback_data="action_buy")
    kb.adjust(3, 3, 1, 1)
    await state.set_state(BuyStone.shape)
    await callback.message.edit_text("💎 *Форма огранки?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.shape, F.data == "shape_custom")
async def buy_step3_shape_custom(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.shape_custom)
    await callback.message.edit_text("✏️ Напиши форму огранки:")


@dp.message(BuyStone.shape_custom)
async def buy_step3_shape_text(message: Message, state: FSMContext):
    await state.update_data(shape=message.text.strip())
    await buy_step4_carat_msg(message, state)


@dp.callback_query(BuyStone.shape, F.data.startswith("shape_"))
async def buy_step3_shape_selected(callback: CallbackQuery, state: FSMContext):
    shape = callback.data.replace("shape_", "")
    await state.update_data(shape=shape)
    await buy_step4_carat_cb(callback, state)


# ============================================================
# КУПИЛИ — шаг 4: каратность
# ============================================================

async def buy_step4_carat_cb(callback: CallbackQuery, state: FSMContext):
    kb = InlineKeyboardBuilder()
    kb.button(text="0.1–0.9 кар", callback_data="carat_hint_small")
    kb.button(text="1–4 кар", callback_data="carat_hint_mid")
    kb.button(text="5+ кар", callback_data="carat_hint_large")
    kb.button(text="✏️ Ввести точно", callback_data="carat_custom")
    kb.adjust(3, 1)
    await state.set_state(BuyStone.carat)
    await callback.message.edit_text("⚖️ *Вес (каратов)?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")

async def buy_step4_carat_msg(message: Message, state: FSMContext):
    kb = InlineKeyboardBuilder()
    kb.button(text="0.1–0.9 кар", callback_data="carat_hint_small")
    kb.button(text="1–4 кар", callback_data="carat_hint_mid")
    kb.button(text="5+ кар", callback_data="carat_hint_large")
    kb.button(text="✏️ Ввести точно", callback_data="carat_custom")
    kb.adjust(3, 1)
    await state.set_state(BuyStone.carat)
    await message.answer("⚖️ *Вес (каратов)?*",
                        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.carat, F.data.in_({"carat_custom", "carat_hint_small", "carat_hint_mid", "carat_hint_large"}))
async def buy_step4_carat_custom(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.carat_custom)
    await callback.message.edit_text("✏️ Введи точный вес в каратах (например: 1.23):")


@dp.message(BuyStone.carat_custom)
async def buy_step4_carat_text(message: Message, state: FSMContext):
    try:
        carat = float(message.text.strip().replace(",", "."))
        await state.update_data(carat=carat)
        await buy_step5_color_msg(message, state)
    except:
        await message.answer("❌ Введи число, например: 1.23")



# ============================================================
# КУПИЛИ — шаг 5: цвет
# ============================================================

async def buy_step5_color_cb(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    stone_type = data.get("stone_type")
    kb = InlineKeyboardBuilder()
    if stone_type == "diamond":
        for c in ["D", "E", "F", "G", "H", "I", "J", "K"]:
            kb.button(text=c, callback_data=f"color_{c}")
        kb.adjust(4)
    else:
        for c in ["AAA", "AA", "A", "B"]:
            kb.button(text=c, callback_data=f"color_{c}")
        kb.adjust(4)
    kb.button(text="✏️ Другой", callback_data="color_custom")
    await state.set_state(BuyStone.color)
    await callback.message.edit_text("🎨 *Цвет?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")

async def buy_step5_color_msg(message: Message, state: FSMContext):
    data = await state.get_data()
    stone_type = data.get("stone_type")
    kb = InlineKeyboardBuilder()
    if stone_type == "diamond":
        for c in ["D", "E", "F", "G", "H", "I", "J", "K"]:
            kb.button(text=c, callback_data=f"color_{c}")
        kb.adjust(4)
    else:
        for c in ["AAA", "AA", "A", "B"]:
            kb.button(text=c, callback_data=f"color_{c}")
        kb.adjust(4)
    kb.button(text="✏️ Другой", callback_data="color_custom")
    await state.set_state(BuyStone.color)
    await message.answer("🎨 *Цвет?*", reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.color, F.data == "color_custom")
async def buy_step5_color_custom(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.color)
    await state.update_data(waiting_color_text=True)
    await callback.message.edit_text("✏️ Введи цвет:")


@dp.callback_query(BuyStone.color, F.data.startswith("color_"))
async def buy_step5_color_selected(callback: CallbackQuery, state: FSMContext):
    color = callback.data.replace("color_", "")
    await state.update_data(color=color)
    await buy_step6_clarity_cb(callback, state)


@dp.message(BuyStone.color)
async def buy_step5_color_text(message: Message, state: FSMContext):
    await state.update_data(color=message.text.strip())
    await buy_step6_clarity_msg(message, state)


# ============================================================
# КУПИЛИ — шаг 6: чистота
# ============================================================

async def buy_step6_clarity_cb(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    stone_type = data.get("stone_type")
    kb = InlineKeyboardBuilder()
    if stone_type == "diamond":
        for c in ["IF", "VVS1", "VVS2", "VS1", "VS2", "SI1", "SI2", "I1"]:
            kb.button(text=c, callback_data=f"clarity_{c}")
        kb.adjust(4)
    else:
        for c in ["AAA", "AA+", "AA", "A+"]:
            kb.button(text=c, callback_data=f"clarity_{c}")
        kb.adjust(4)
    kb.button(text="✏️ Другая", callback_data="clarity_custom")
    await state.set_state(BuyStone.clarity)
    await callback.message.edit_text("🔍 *Чистота?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")

async def buy_step6_clarity_msg(message: Message, state: FSMContext):
    data = await state.get_data()
    stone_type = data.get("stone_type")
    kb = InlineKeyboardBuilder()
    if stone_type == "diamond":
        for c in ["IF", "VVS1", "VVS2", "VS1", "VS2", "SI1", "SI2", "I1"]:
            kb.button(text=c, callback_data=f"clarity_{c}")
        kb.adjust(4)
    else:
        for c in ["AAA", "AA+", "AA", "A+"]:
            kb.button(text=c, callback_data=f"clarity_{c}")
        kb.adjust(4)
    kb.button(text="✏️ Другая", callback_data="clarity_custom")
    await state.set_state(BuyStone.clarity)
    await message.answer("🔍 *Чистота?*", reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.clarity, F.data == "clarity_custom")
async def buy_step6_clarity_custom(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("✏️ Введи чистоту:")


@dp.callback_query(BuyStone.clarity, F.data.startswith("clarity_"))
async def buy_step6_clarity_selected(callback: CallbackQuery, state: FSMContext):
    clarity = callback.data.replace("clarity_", "")
    await state.update_data(clarity=clarity)
    await buy_step7_price_cb(callback, state)


@dp.message(BuyStone.clarity)
async def buy_step6_clarity_text(message: Message, state: FSMContext):
    await state.update_data(clarity=message.text.strip())
    await buy_step7_price_msg(message, state)


# ============================================================
# КУПИЛИ — шаг 7: цена
# ============================================================

async def buy_step7_price_cb(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    is_natural = data.get("origin") == "Природный"
    prompt = ("💵 *Цена за карат?*\n\nВведи сумму (например: 800):" if is_natural
              else "💵 *Цена за камень?*\n\nВведи сумму (например: 5000):")
    await state.set_state(BuyStone.price)
    await callback.message.edit_text(prompt, parse_mode="Markdown")

async def buy_step7_price_msg(message: Message, state: FSMContext):
    data = await state.get_data()
    is_natural = data.get("origin") == "Природный"
    prompt = ("💵 *Цена за карат?*\n\nВведи сумму (например: 800):" if is_natural
              else "💵 *Цена за камень?*\n\nВведи сумму (например: 5000):")
    await state.set_state(BuyStone.price)
    await message.answer(prompt, parse_mode="Markdown")


@dp.message(BuyStone.price)
async def buy_step7_price_entered(message: Message, state: FSMContext):
    try:
        val = float(message.text.strip().replace(",", ".").replace(" ", ""))
        data = await state.get_data()
        if data.get("origin") == "Природный":
            await state.update_data(price_per_carat=val)
        else:
            await state.update_data(price=val)
        kb = InlineKeyboardBuilder()
        for cur in ["CNY", "USD", "THB", "RUB"]:
            kb.button(text=cur, callback_data=f"currency_{cur}")
        kb.adjust(4)
        await state.set_state(BuyStone.currency)
        await message.answer("💱 *Валюта?*", reply_markup=kb.as_markup(), parse_mode="Markdown")
    except:
        await message.answer("❌ Введи число, например: 800")


# ============================================================
# КУПИЛИ — шаг 8: валюта → курс → поставщик
# ============================================================

async def _show_price_summary(target, state: FSMContext):
    data = await state.get_data()
    origin = data.get("origin", "")
    currency = data.get("currency", "USD")
    rate = data.get("exchange_rate", 1.0) or 1.0

    if origin == "Природный":
        carat = data.get("carat", 0)
        ppc = data.get("price_per_carat", 0)
        total = round(carat * ppc, 2)
        await state.update_data(price=total)
        carat_str = f"{carat:g}"
        if currency == "USD":
            summary = f"{carat_str} × {ppc:,.0f} USD = *${total:,.0f}*"
        else:
            total_usd = round(total / rate, 2)
            summary = (f"{carat_str} × {ppc:,.0f} {currency} = "
                       f"{total:,.0f} {currency} = *${total_usd:,.0f} USD*")
    else:
        price = data.get("price", 0)
        if currency == "USD":
            summary = f"*${price:,.0f} USD*"
        else:
            price_usd = round(price / rate, 2)
            summary = f"{price:,.0f} {currency} = *${price_usd:,.0f} USD*"

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Верно", callback_data="price_ok")
    kb.button(text="✏️ Изменить", callback_data="price_edit")
    kb.adjust(2)
    await state.set_state(BuyStone.price_confirm)
    text = f"💵 *Итого:*\n\n{summary}"
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="Markdown")
    else:
        await target.answer(text, reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(BuyStone.price_confirm, F.data == "price_ok")
async def buy_price_confirmed(callback: CallbackQuery, state: FSMContext):
    await _show_buy_supplier(callback, state)


@dp.callback_query(BuyStone.price_confirm, F.data == "price_edit")
async def buy_price_edit(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    is_natural = data.get("origin") == "Природный"
    prompt = ("💵 *Цена за карат?*\n\nВведи сумму (например: 800):" if is_natural
              else "💵 *Цена за камень?*\n\nВведи сумму (например: 5000):")
    await state.set_state(BuyStone.price)
    await callback.message.edit_text(prompt, parse_mode="Markdown")


async def _show_buy_supplier(target, state: FSMContext):
    res = supabase.table("counterparties").select("id,name").eq("type", "supplier").execute()
    suppliers = res.data or []
    kb = InlineKeyboardBuilder()
    for s in suppliers:
        kb.button(text=s["name"], callback_data=f"supplier_{s['id']}")
    kb.button(text="✏️ Новый поставщик", callback_data="supplier_new")
    kb.adjust(1)
    await state.set_state(BuyStone.supplier)
    if isinstance(target, CallbackQuery):
        await target.message.edit_text("🏪 *Поставщик?*", reply_markup=kb.as_markup(), parse_mode="Markdown")
    else:
        await target.answer("🏪 *Поставщик?*", reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(BuyStone.currency, F.data.startswith("currency_"))
async def buy_step8_currency(callback: CallbackQuery, state: FSMContext):
    currency = callback.data.replace("currency_", "")
    await state.update_data(currency=currency)
    if currency == "USD":
        await state.update_data(exchange_rate=1.0)
        await _show_price_summary(callback, state)
    else:
        await state.set_state(BuyStone.exchange_rate)
        await callback.message.edit_text(
            f"💱 *Курс к USD сегодня?*\n1 USD = ? {currency}\n\nВведи число (например: 7.25):",
            parse_mode="Markdown")

@dp.message(BuyStone.exchange_rate)
async def buy_step8_rate(message: Message, state: FSMContext):
    try:
        rate = float(message.text.strip().replace(",", "."))
        if rate <= 0:
            raise ValueError
        await state.update_data(exchange_rate=rate)
        await _show_price_summary(message, state)
    except:
        await message.answer("❌ Введи число больше нуля, например: 7.25")


@dp.callback_query(BuyStone.supplier, F.data == "supplier_new")
async def buy_step8_supplier_new(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BuyStone.new_supplier)
    await callback.message.edit_text("✏️ Введи имя нового поставщика:")


@dp.message(BuyStone.new_supplier)
async def buy_step8_supplier_new_name(message: Message, state: FSMContext):
    name = message.text.strip()
    try:
        res = supabase.table("counterparties").insert({"name": name, "type": "supplier"}).execute()
        supplier_id = res.data[0]["id"]
        await state.update_data(supplier_id=supplier_id, supplier_name=name)
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        return
    await _ask_photo_before_confirm(message, state)


@dp.callback_query(BuyStone.supplier, F.data.startswith("supplier_"))
async def buy_step8_supplier_selected(callback: CallbackQuery, state: FSMContext):
    supplier_id = callback.data.replace("supplier_", "")
    res = supabase.table("counterparties").select("name").eq("id", supplier_id).execute()
    supplier_name = res.data[0]["name"] if res.data else "—"
    await state.update_data(supplier_id=supplier_id, supplier_name=supplier_name)
    await _ask_photo_before_confirm(callback, state)


# ============================================================
# КУПИЛИ — шаг 9: подтверждение
# ============================================================

def _fmt_price_with_usd(data: dict) -> str:
    price = data.get("price", 0) or 0
    currency = data.get("currency", "USD")
    rate = data.get("exchange_rate", 1.0) or 1.0
    price_usd = price / rate
    s = f"{price:,.0f} {currency}"
    if currency != "USD":
        s += f" = {price_usd:,.0f} USD"
    return s

async def _show_buy_confirm(target, state: FSMContext):
    data = await state.get_data()
    type_name = {"diamond": "Бриллиант", "emerald": "Изумруд",
                 "ruby": "Рубин", "spinel": "Шпинель"}.get(data.get("stone_type"), data.get("stone_type"))
    media = data.get("media") or []
    if media:
        n_photo = sum(1 for m in media if m["type"] == "photo")
        n_video = sum(1 for m in media if m["type"] == "video")
        n_anim  = sum(1 for m in media if m["type"] == "animation")
        parts = ([f"{n_photo} фото"] if n_photo else []) + \
                ([f"{n_video} видео"] if n_video else []) + \
                ([f"{n_anim} GIF"] if n_anim else [])
        media_status = f"✅ {len(media)} шт ({', '.join(parts)})"
    else:
        media_status = "—"
    cert_status = "✅ Есть" if data.get("cert_file_id") else "—"
    text = (
        f"✅ *Проверь данные:*\n\n"
        f"Тип: {type_name} ({data.get('origin', '—')})\n"
        f"Форма: {data.get('shape', '—')}\n"
        f"Вес: {data.get('carat', '—')} кар\n"
        f"Цвет: {data.get('color', '—')}\n"
        f"Чистота: {data.get('clarity', '—')}\n"
        f"Цена: {_fmt_price_with_usd(data)}\n"
        f"Поставщик: {data.get('supplier_name', '—')}\n"
        f"Медиа: {media_status}\n"
        f"Сертификат: {cert_status}"
    )
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Внести в базу", callback_data="buy_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(BuyStone.confirm)
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="Markdown")
    else:
        await target.answer(text, reply_markup=kb.as_markup(), parse_mode="Markdown")


def _skip_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="⏭ Пропустить", callback_data="skip_media")
    return kb.as_markup()


async def _ask_photo(target, stone_code: str, price_usd: float):
    text = (f"✅ *Камень внесён!*\n\nКод: `{stone_code}`\n"
            f"Стоимость: {price_usd:,.0f} USD\n\n"
            f"📸 Загрузи фото камня (или нажми Пропустить)")
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=_skip_kb(), parse_mode="Markdown")
    else:
        await target.answer(text, reply_markup=_skip_kb(), parse_mode="Markdown")


async def _ask_certificate(target):
    text = "📄 Есть сертификат? Загрузи фото или нажми Пропустить"
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=_skip_kb(), parse_mode="Markdown")
    else:
        await target.answer(text, reply_markup=_skip_kb(), parse_mode="Markdown")


async def _ask_photo_before_confirm(target, state: FSMContext):
    await state.update_data(media=[], media_msg_id=None)
    await state.set_state(BuyStone.photo)
    text = "📸 Загрузи фото/видео камня (до 4 штук)\nМожно отправить пачкой или по одному:"
    kb = InlineKeyboardBuilder()
    kb.button(text="⏭ Пропустить", callback_data="skip_media")
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=kb.as_markup())
    else:
        await target.answer(text, reply_markup=kb.as_markup())


async def _ask_cert_before_confirm(target, state: FSMContext):
    await state.set_state(BuyStone.certificate)
    text = "📄 Есть сертификат? Загрузи фото или нажми Пропустить:"
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=_skip_kb())
    else:
        await target.answer(text, reply_markup=_skip_kb())


async def _finish(target, state: FSMContext):
    await state.clear()
    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    text = "✅ Готово!"
    if isinstance(target, CallbackQuery):
        await target.message.edit_text(text, reply_markup=kb.as_markup())
    else:
        await target.answer(text, reply_markup=kb.as_markup())


async def _insert_stone(callback: CallbackQuery, state: FSMContext, data: dict):
    user_id = get_user_id(callback.from_user.id)
    stone_code = next_stone_code()

    currency = data.get("currency", "USD")
    price = data.get("price", 0)
    rate = data.get("exchange_rate", 1.0) or 1.0
    price_usd = round(price / rate, 2)

    origin_str = data.get("origin", "")
    stone_type = data.get("stone_type", "")
    full_type = f"{stone_type} ({origin_str})" if origin_str else stone_type

    try:
        res = supabase.table("stones").insert({
            "stone_code": stone_code,
            "stone_type": full_type,
            "carat": data.get("carat"),
            "shape": data.get("shape"),
            "color": data.get("color"),
            "clarity": data.get("clarity"),
            "purchase_date": str(date.today()),
            "purchase_price": price,
            "purchase_currency": currency,
            "purchase_price_usd": price_usd,
            "exchange_rate": rate,
            "supplier_id": data.get("supplier_id"),
            "status": "in_stock",
            "created_by": user_id,
        }).execute()

        stone_id = res.data[0]["id"]

        supabase.table("operations").insert({
            "operation_type": "purchase_stone",
            "entity_type": "stone",
            "entity_id": stone_id,
            "counterparty_id": data.get("supplier_id"),
            "amount": price,
            "currency": currency,
            "amount_usd": price_usd,
            "exchange_rate": rate,
            "created_by": user_id,
        }).execute()

        for m in (data.get("media") or []):
            supabase.table("media_files").insert({
                "entity_type": "stone", "entity_id": stone_id,
                "file_type": m["type"], "file_url": m["file_id"],
            }).execute()

        if data.get("cert_file_id"):
            supabase.table("media_files").insert({
                "entity_type": "stone", "entity_id": stone_id,
                "file_type": "certificate_scan", "file_url": data["cert_file_id"],
            }).execute()
            supabase.table("certificates").insert({
                "stone_id": stone_id, "laboratory": "pending", "cert_number": "pending",
            }).execute()

        await state.clear()
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text(
            f"✅ *Камень внесён!*\n\nКод: `{stone_code}`\nСтоимость: {price_usd:,.0f} USD",
            reply_markup=kb.as_markup(), parse_mode="Markdown")
    except Exception as e:
        await callback.message.edit_text(f"❌ Ошибка: {e}")


# ============================================================
# КУПИЛИ — медиа до 4 файлов (фото/видео/gif)
# ============================================================

@dp.message(BuyStone.photo, F.photo | F.video | F.animation)
async def buy_media_received(message: Message, state: FSMContext):
    if message.animation:
        file_id, media_type = message.animation.file_id, "animation"
    elif message.video:
        file_id, media_type = message.video.file_id, "video"
    else:
        file_id, media_type = message.photo[-1].file_id, "photo"

    uid = message.from_user.id
    if uid not in _user_media_locks:
        _user_media_locks[uid] = asyncio.Lock()

    async with _user_media_locks[uid]:
        data = await state.get_data()
        media: list = list(data.get("media") or [])
        if len(media) >= 4:
            return
        media.append({"type": media_type, "file_id": file_id})
        n = len(media)
        media_msg_id: int | None = data.get("media_msg_id")

        kb = InlineKeyboardBuilder()
        kb.button(text="✅ Готово", callback_data="media_done")
        kb.button(text="⏭ Пропустить", callback_data="skip_media")
        kb.adjust(2)

        if n >= 4:
            text = "📸 *Добавлено медиа: 4 из 4* — лимит достигнут"
            if media_msg_id:
                try:
                    await bot.edit_message_text(
                        text, chat_id=message.chat.id,
                        message_id=media_msg_id, parse_mode="Markdown")
                except Exception:
                    await message.answer(text, parse_mode="Markdown")
            else:
                await message.answer(text, parse_mode="Markdown")
            await state.update_data(media=media)
        else:
            text = (f"📸 *Добавлено медиа: {n} из 4*\n\n"
                    f"Можно добавить ещё или нажать Готово:")
            if media_msg_id:
                try:
                    await bot.edit_message_text(
                        text, chat_id=message.chat.id, message_id=media_msg_id,
                        reply_markup=kb.as_markup(), parse_mode="Markdown")
                    await state.update_data(media=media)
                except Exception:
                    sent = await message.answer(
                        text, reply_markup=kb.as_markup(), parse_mode="Markdown")
                    await state.update_data(media=media, media_msg_id=sent.message_id)
            else:
                sent = await message.answer(
                    text, reply_markup=kb.as_markup(), parse_mode="Markdown")
                await state.update_data(media=media, media_msg_id=sent.message_id)

    if n >= 4:
        await _ask_cert_before_confirm(message, state)


@dp.callback_query(BuyStone.photo, F.data == "media_done")
async def buy_media_done(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await _ask_cert_before_confirm(callback, state)


@dp.callback_query(BuyStone.photo, F.data == "skip_media")
async def buy_photo_skip(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.update_data(media=[], media_msg_id=None)
    await _ask_cert_before_confirm(callback, state)


# ============================================================
# КУПИЛИ — сертификат
# ============================================================

@dp.message(BuyStone.certificate, F.photo)
async def buy_cert_received(message: Message, state: FSMContext):
    await state.update_data(cert_file_id=message.photo[-1].file_id)
    await _show_buy_confirm(message, state)


@dp.callback_query(BuyStone.certificate, F.data == "skip_media")
async def buy_cert_skip(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await _show_buy_confirm(callback, state)


@dp.callback_query(BuyStone.confirm, F.data == "buy_confirm_yes")
async def buy_save(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()

    origin_str = data.get("origin", "")
    stone_type = data.get("stone_type", "")
    full_type = f"{stone_type} ({origin_str})" if origin_str else stone_type

    dup = supabase.table("stones").select("id").eq("stone_type", full_type) \
        .eq("carat", data.get("carat")).eq("color", data.get("color")) \
        .eq("clarity", data.get("clarity")).eq("supplier_id", data.get("supplier_id")) \
        .execute()

    if dup.data:
        kb = InlineKeyboardBuilder()
        kb.button(text="✅ Да, внести", callback_data="buy_force_yes")
        kb.button(text="❌ Отмена", callback_data="back_menu")
        kb.adjust(1)
        await state.set_state(BuyStone.duplicate_confirm)
        await callback.message.edit_text(
            "⚠️ *Похожий камень уже есть в базе.*\n\nЭто пара? Внести всё равно?",
            reply_markup=kb.as_markup(), parse_mode="Markdown")
        return

    await _insert_stone(callback, state, data)


@dp.callback_query(BuyStone.duplicate_confirm, F.data == "buy_force_yes")
async def buy_save_forced(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await _insert_stone(callback, state, data)


# ============================================================
# ПРОДАЛИ
# ============================================================

@dp.callback_query(F.data == "action_sell")
async def sell_step1(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    data = supabase.table("stones").select("id,stone_code,stone_type,carat") \
        .eq("status", "in_stock").order("created_at", desc=True).execute()

    if not data.data:
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text("Нет камней в наличии.", reply_markup=kb.as_markup())
        return

    kb = InlineKeyboardBuilder()
    for s in data.data:
        kb.button(text=fmt_stone_btn(s),
                 callback_data=f"sell_stone_{s['id']}")
    kb.button(text="◀️ Назад", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(SellStone.select_stone)
    await callback.message.edit_text("💰 *Какой камень продали?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(SellStone.select_stone, F.data.startswith("sell_stone_"))
async def sell_step2_price(callback: CallbackQuery, state: FSMContext):
    stone_id = callback.data.replace("sell_stone_", "")
    res = supabase.table("stones").select("stone_code,stone_type,carat").eq("id", stone_id).execute()
    stone = res.data[0]
    await state.update_data(stone_id=stone_id, stone_code=stone["stone_code"])
    await state.set_state(SellStone.price)
    await callback.message.edit_text(
        f"💰 *{abbr_type(stone['stone_type'])}* {stone['carat']}кар · `{abbr_code(stone['stone_code'])}`\n\n"
        f"Введи цену продажи:", parse_mode="Markdown")


@dp.message(SellStone.price)
async def sell_step3_currency(message: Message, state: FSMContext):
    try:
        price = float(message.text.strip().replace(",", ".").replace(" ", ""))
        await state.update_data(price=price)
        kb = InlineKeyboardBuilder()
        for cur in ["CNY", "USD", "THB", "RUB"]:
            kb.button(text=cur, callback_data=f"sell_cur_{cur}")
        kb.adjust(4)
        await state.set_state(SellStone.currency)
        await message.answer("💱 *Валюта?*", reply_markup=kb.as_markup(), parse_mode="Markdown")
    except:
        await message.answer("❌ Введи число")


async def _show_sell_client(target, state: FSMContext):
    res = supabase.table("counterparties").select("id,name").eq("type", "client").execute()
    clients = res.data or []
    kb = InlineKeyboardBuilder()
    for c in clients:
        kb.button(text=c["name"], callback_data=f"sell_client_{c['id']}")
    kb.button(text="✏️ Новый клиент", callback_data="sell_client_new")
    kb.adjust(1)
    await state.set_state(SellStone.client)
    if isinstance(target, CallbackQuery):
        await target.message.edit_text("👤 *Кому продали?*", reply_markup=kb.as_markup(), parse_mode="Markdown")
    else:
        await target.answer("👤 *Кому продали?*", reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(SellStone.currency, F.data.startswith("sell_cur_"))
async def sell_step4_currency(callback: CallbackQuery, state: FSMContext):
    currency = callback.data.replace("sell_cur_", "")
    await state.update_data(currency=currency)
    if currency == "USD":
        await state.update_data(exchange_rate=1.0)
        await _show_sell_client(callback, state)
    else:
        await state.set_state(SellStone.exchange_rate)
        await callback.message.edit_text(
            f"💱 *Курс на момент сделки:*\n1 USD = ? {currency}\n\nВведи число (например: 7.25):",
            parse_mode="Markdown")

@dp.message(SellStone.exchange_rate)
async def sell_step4_rate(message: Message, state: FSMContext):
    try:
        rate = float(message.text.strip().replace(",", "."))
        if rate <= 0:
            raise ValueError
        await state.update_data(exchange_rate=rate)
        await _show_sell_client(message, state)
    except:
        await message.answer("❌ Введи число больше нуля, например: 7.25")


@dp.callback_query(SellStone.client, F.data == "sell_client_new")
async def sell_client_new(callback: CallbackQuery, state: FSMContext):
    await state.set_state(SellStone.new_client)
    await callback.message.edit_text("✏️ Введи имя нового клиента:")


@dp.message(SellStone.new_client)
async def sell_client_new_name(message: Message, state: FSMContext):
    name = message.text.strip()
    try:
        res = supabase.table("counterparties").insert({"name": name, "type": "client"}).execute()
        client_id = res.data[0]["id"]
        await state.update_data(client_id=client_id, client_name=name)
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        return
    data = await state.get_data()
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить продажу", callback_data="sell_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(SellStone.confirm)
    await message.answer(
        f"✅ *Продажа:*\n\nКамень: {data['stone_code']}\n"
        f"Цена: {_fmt_price_with_usd(data)}\nКлиент: {name}",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(SellStone.client, F.data.startswith("sell_client_"))
async def sell_confirm(callback: CallbackQuery, state: FSMContext):
    client_id = callback.data.replace("sell_client_", "")
    res = supabase.table("counterparties").select("name").eq("id", client_id).execute()
    client_name = res.data[0]["name"] if res.data else "—"

    await state.update_data(client_id=client_id, client_name=client_name)
    data = await state.get_data()

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить продажу", callback_data="sell_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(SellStone.confirm)
    await callback.message.edit_text(
        f"✅ *Продажа:*\n\nКамень: {data['stone_code']}\n"
        f"Цена: {_fmt_price_with_usd(data)}\nКлиент: {client_name}",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(SellStone.confirm, F.data == "sell_confirm_yes")
async def sell_save(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = get_user_id(callback.from_user.id)

    rate = data.get("exchange_rate", 1.0) or 1.0
    amount_usd = round(data["price"] / rate, 2)

    try:
        supabase.table("stones").update({"status": "sold"}).eq("id", data["stone_id"]).execute()
        supabase.table("operations").insert({
            "operation_type": "sale_stone",
            "entity_type": "stone",
            "entity_id": data["stone_id"],
            "counterparty_id": data.get("client_id"),
            "amount": data["price"],
            "currency": data["currency"],
            "amount_usd": amount_usd,
            "exchange_rate": rate,
            "created_by": user_id,
        }).execute()

        await state.clear()
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text(
            f"✅ *Продажа записана!*\n\n{data['stone_code']} продан\n"
            f"Сумма: {amount_usd:,.0f} USD",
            reply_markup=kb.as_markup(), parse_mode="Markdown")
    except Exception as e:
        await callback.message.edit_text(f"❌ Ошибка: {e}")


# ============================================================
# ЮВЕЛИРУ (передача)
# ============================================================

@dp.callback_query(F.data == "action_transfer")
async def transfer_step1(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    data = supabase.table("stones").select("id,stone_code,stone_type,carat") \
        .eq("status", "in_stock").order("created_at", desc=True).execute()

    if not data.data:
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text("Нет камней в наличии.", reply_markup=kb.as_markup())
        return

    kb = InlineKeyboardBuilder()
    for s in data.data:
        kb.button(text=fmt_stone_btn(s),
                 callback_data=f"transfer_stone_{s['id']}")
    kb.button(text="◀️ Назад", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(TransferStone.select_stone)
    await callback.message.edit_text("📤 *Какой камень отправляем ювелиру?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(TransferStone.select_stone, F.data.startswith("transfer_stone_"))
async def transfer_step2_partner(callback: CallbackQuery, state: FSMContext):
    stone_id = callback.data.replace("transfer_stone_", "")
    res = supabase.table("stones").select("stone_code").eq("id", stone_id).execute()
    stone_code = res.data[0]["stone_code"]
    await state.update_data(stone_id=stone_id, stone_code=stone_code)

    partners = supabase.table("counterparties").select("id,name") \
        .in_("type", ["partner"]).execute()

    kb = InlineKeyboardBuilder()
    for p in (partners.data or []):
        kb.button(text=p["name"], callback_data=f"transfer_partner_{p['id']}")
    kb.button(text="✏️ Новый", callback_data="transfer_partner_new")
    kb.button(text="◀️ Назад", callback_data="action_transfer")
    kb.adjust(1)
    await state.set_state(TransferStone.partner)
    await callback.message.edit_text("👤 *Кому передаём?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(TransferStone.partner, F.data == "transfer_partner_new")
async def transfer_partner_new(callback: CallbackQuery, state: FSMContext):
    await state.set_state(TransferStone.new_partner)
    await callback.message.edit_text("✏️ Введи имя нового партнёра:")


@dp.message(TransferStone.new_partner)
async def transfer_partner_new_name(message: Message, state: FSMContext):
    name = message.text.strip()
    try:
        res = supabase.table("counterparties").insert({"name": name, "type": "partner"}).execute()
        partner_id = res.data[0]["id"]
        await state.update_data(partner_id=partner_id, partner_name=name)
    except Exception as e:
        await message.answer(f"❌ Ошибка сохранения: {e}")
        return
    data = await state.get_data()
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить", callback_data="transfer_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(TransferStone.confirm)
    await message.answer(
        f"📤 *Передача ювелиру:*\n\nКамень: {data['stone_code']}\nКому: {name}",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(TransferStone.partner, F.data.startswith("transfer_partner_"))
async def transfer_confirm(callback: CallbackQuery, state: FSMContext):
    partner_id = callback.data.replace("transfer_partner_", "")
    res = supabase.table("counterparties").select("name").eq("id", partner_id).execute()
    partner_name = res.data[0]["name"] if res.data else "—"

    await state.update_data(partner_id=partner_id, partner_name=partner_name)
    data = await state.get_data()

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить", callback_data="transfer_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(TransferStone.confirm)
    await callback.message.edit_text(
        f"📤 *Передача ювелиру:*\n\nКамень: {data['stone_code']}\nКому: {partner_name}",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(TransferStone.confirm, F.data == "transfer_confirm_yes")
async def transfer_save(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = get_user_id(callback.from_user.id)

    try:
        supabase.table("stones").update({"status": "at_partner"}).eq("id", data["stone_id"]).execute()
        supabase.table("operations").insert({
            "operation_type": "transfer_to_partner",
            "entity_type": "stone",
            "entity_id": data["stone_id"],
            "counterparty_id": data.get("partner_id"),
            "created_by": user_id,
        }).execute()

        await state.clear()
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text(
            f"✅ *Передача записана!*\n\n{data['stone_code']} → {data['partner_name']}",
            reply_markup=kb.as_markup(), parse_mode="Markdown")
    except Exception as e:
        await callback.message.edit_text(f"❌ Ошибка: {e}")


# ============================================================
# ВОЗВРАТ
# ============================================================

_RETURN_REASONS = [
    ("Не подошёл",  "not_fit"),
    ("Дефект",      "defect"),
    ("Другое",      "other"),
]

@dp.callback_query(F.data == "action_return")
async def return_step1(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    data = supabase.table("stones").select("id,stone_code,stone_type,carat,status") \
        .in_("status", ["sold", "at_partner"]).order("created_at", desc=True).execute()

    if not data.data:
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text(
            "Нет камней для возврата (нет проданных или у партнёра).",
            reply_markup=kb.as_markup())
        return

    kb = InlineKeyboardBuilder()
    for s in data.data:
        status_label = "продан" if s["status"] == "sold" else "у партнёра"
        btn_text = f"{fmt_stone_btn(s)}  [{status_label}]"
        kb.button(text=btn_text, callback_data=f"ret_stone_{s['id']}")
    kb.button(text="◀️ Назад", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(ReturnStone.select_stone)
    await callback.message.edit_text("↩️ *Возврат — выбери камень:*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(ReturnStone.select_stone, F.data.startswith("ret_stone_"))
async def return_step2_counterparty(callback: CallbackQuery, state: FSMContext):
    stone_id = callback.data.replace("ret_stone_", "")
    res = supabase.table("stones").select("stone_code,stone_type,carat,status").eq("id", stone_id).execute()
    stone = res.data[0]
    await state.update_data(stone_id=stone_id, stone_code=stone["stone_code"],
                            original_status=stone["status"])

    cp_type = "client" if stone["status"] == "sold" else "partner"
    cp_label = "клиента" if cp_type == "client" else "партнёра"
    counterparties = supabase.table("counterparties").select("id,name") \
        .eq("type", cp_type).execute().data or []

    kb = InlineKeyboardBuilder()
    for c in counterparties:
        kb.button(text=c["name"], callback_data=f"ret_cp_{c['id']}")
    kb.button(text="◀️ Назад", callback_data="action_return")
    kb.adjust(1)
    await state.set_state(ReturnStone.counterparty)
    await callback.message.edit_text(
        f"↩️ *Возврат камня* `{stone['stone_code']}`\n\nОт кого возврат ({cp_label})?",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(ReturnStone.counterparty, F.data.startswith("ret_cp_"))
async def return_step3_reason(callback: CallbackQuery, state: FSMContext):
    cp_id = callback.data.replace("ret_cp_", "")
    res = supabase.table("counterparties").select("name").eq("id", cp_id).execute()
    cp_name = res.data[0]["name"] if res.data else "—"
    await state.update_data(counterparty_id=cp_id, counterparty_name=cp_name)

    kb = InlineKeyboardBuilder()
    for label, key in _RETURN_REASONS:
        kb.button(text=label, callback_data=f"ret_reason_{key}")
    kb.button(text="◀️ Назад", callback_data="action_return")
    kb.adjust(1)
    await state.set_state(ReturnStone.reason)
    await callback.message.edit_text("↩️ *Причина возврата?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(ReturnStone.reason, F.data.startswith("ret_reason_"))
async def return_step4_confirm(callback: CallbackQuery, state: FSMContext):
    reason_key = callback.data.replace("ret_reason_", "")
    reason_label = next((lbl for lbl, k in _RETURN_REASONS if k == reason_key), reason_key)
    await state.update_data(reason=reason_key, reason_label=reason_label)
    data = await state.get_data()

    op_type = "return_from_client" if data["original_status"] == "sold" else "return_from_partner"
    await state.update_data(op_type=op_type)

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить", callback_data="ret_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="back_menu")
    kb.adjust(1)
    await state.set_state(ReturnStone.confirm)
    await callback.message.edit_text(
        f"↩️ *Подтверди возврат:*\n\n"
        f"Камень: `{data['stone_code']}`\n"
        f"От: {data['counterparty_name']}\n"
        f"Причина: {reason_label}\n"
        f"Новый статус: в наличии",
        reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(ReturnStone.confirm, F.data == "ret_confirm_yes")
async def return_save(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = get_user_id(callback.from_user.id)

    try:
        supabase.table("stones").update({"status": "in_stock"}) \
            .eq("id", data["stone_id"]).execute()
        supabase.table("operations").insert({
            "operation_type": data["op_type"],
            "entity_type":    "stone",
            "entity_id":      data["stone_id"],
            "counterparty_id": data.get("counterparty_id"),
            "notes":           data.get("reason_label"),
            "created_by":      user_id,
        }).execute()

        await state.clear()
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Меню", callback_data="back_menu")
        await callback.message.edit_text(
            f"✅ *Возврат записан!*\n\n"
            f"{data['stone_code']} → в наличии\n"
            f"От: {data['counterparty_name']}\n"
            f"Причина: {data['reason_label']}",
            reply_markup=kb.as_markup(), parse_mode="Markdown")
    except Exception as e:
        await callback.message.edit_text(f"❌ Ошибка: {e}")


# ============================================================
# Назад в меню
# ============================================================

@dp.callback_query(F.data == "back_menu")
async def back_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("💎 *Jewelry AI*\n\nВыбери действие:",
                                     reply_markup=main_keyboard(), parse_mode="Markdown")


# ============================================================
# КАРТОЧКА КАМНЯ — хелпер + кнопочный флоу + /stone
# ============================================================

async def _send_stone_card(msg: Message, s: dict):
    """Порядок: медиа альбом → сертификат → текст карточки с кнопкой."""
    supplier_name = "—"
    if s.get("supplier_id"):
        sp = supabase.table("counterparties").select("name").eq("id", s["supplier_id"]).execute()
        if sp.data:
            supplier_name = sp.data[0]["name"]

    creator_name = "—"
    if s.get("created_by"):
        u = supabase.table("users").select("name").eq("id", s["created_by"]).execute()
        if u.data:
            creator_name = u.data[0]["name"]
    creator_date = s["created_at"][:10] if s.get("created_at") else "—"

    color = abbr_color(s.get("color") or "")
    clarity = s.get("clarity") or ""
    color_clarity = " · ".join(filter(None, [color, clarity]))

    currency = s.get("purchase_currency") or "USD"
    price = s.get("purchase_price", 0) or 0
    price_usd = s.get("purchase_price_usd", 0) or 0
    if currency == "USD":
        price_line = f"Цена: {price:,.0f} USD"
    else:
        price_line = f"Цена: {price:,.0f} {currency} ({price_usd:,.0f} USD)"

    lines = [
        f"💎 *{abbr_type(s['stone_type'])}* {s['carat']}кар · `{abbr_code(s['stone_code'])}`",
        f"Форма: {s.get('shape') or '—'}",
    ]
    if color_clarity:
        lines.append(f"Цвет/Чистота: {color_clarity}")
    lines += [
        f"Статус: {get_status_emoji(s['status'])} {get_status_name(s['status'])}",
        "",
        f"Поставщик: {supplier_name}",
        price_line,
        "",
        f"Внёс: {creator_name} · {creator_date}",
    ]

    if s["status"] == "sold":
        sale = supabase.table("operations").select(
            "created_by,created_at,counterparty_id"
        ).eq("entity_id", s["id"]).eq("operation_type", "sale_stone") \
         .order("created_at", desc=True).limit(1).execute()
        if sale.data:
            op = sale.data[0]
            seller_name = "—"
            if op.get("created_by"):
                su = supabase.table("users").select("name").eq("id", op["created_by"]).execute()
                if su.data:
                    seller_name = su.data[0]["name"]
            sale_date = op["created_at"][:10] if op.get("created_at") else "—"
            client_name = "—"
            if op.get("counterparty_id"):
                cl = supabase.table("counterparties").select("name").eq("id", op["counterparty_id"]).execute()
                if cl.data:
                    client_name = cl.data[0]["name"]
            lines.append(f"Продал: {seller_name} · {sale_date} → {client_name}")

    media_rows = supabase.table("media_files") \
        .select("file_type,file_url") \
        .eq("entity_type", "stone").eq("entity_id", s["id"]) \
        .execute().data or []

    visuals = [r for r in media_rows if r["file_type"] in ("photo", "video", "animation")]
    cert = next((r for r in media_rows if r["file_type"] == "certificate_scan"), None)

    if not visuals and not cert:
        lines += ["", "📷 Нет загруженных фото"]

    # Сначала медиа, потом карточка с кнопкой
    if visuals:
        media_group = []
        for r in visuals[:10]:
            ft, fid = r["file_type"], r["file_url"]
            if ft == "photo":
                media_group.append(InputMediaPhoto(media=fid))
            elif ft == "video":
                media_group.append(InputMediaVideo(media=fid))
            else:
                media_group.append(InputMediaAnimation(media=fid))
        await msg.answer_media_group(media_group)

    if cert:
        await msg.answer_photo(cert["file_url"], caption="📜 Сертификат")

    kb = InlineKeyboardBuilder()
    kb.button(text="◀️ Меню", callback_data="back_menu")
    await msg.answer("\n".join(lines), reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(F.data == "action_view")
async def view_filter_select(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.answer()
    kb = InlineKeyboardBuilder()
    kb.button(text="🌿 Природные",  callback_data="view_filter_natural")
    kb.button(text="⚗️ Синтетика", callback_data="view_filter_synthetic")
    kb.button(text="📋 Все",        callback_data="view_filter_all")
    kb.button(text="◀️ Меню",       callback_data="back_menu")
    kb.adjust(2, 1, 1)
    await callback.message.edit_text("👁 *Какие камни показать?*",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


async def _view_show_list(callback: CallbackQuery, filter_type: str):
    q = supabase.table("stones").select("id,stone_code,stone_type,carat") \
        .not_.in_("status", ["sold", "written_off"]) \
        .order("created_at", desc=True)
    if filter_type == "natural":
        q = q.ilike("stone_type", "%(Природный)%")
    elif filter_type == "synthetic":
        q = q.ilike("stone_type", "%(Синтетический)%")
    data = q.execute()

    if not data.data:
        kb = InlineKeyboardBuilder()
        kb.button(text="◀️ Назад", callback_data="action_view")
        await callback.message.edit_text("Нет камней по этому фильтру.",
                                         reply_markup=kb.as_markup())
        return

    stone_ids = [s["id"] for s in data.data]
    media_counts: dict = {}
    try:
        media_rows = supabase.table("media_files") \
            .select("entity_id,file_type") \
            .eq("entity_type", "stone") \
            .in_("entity_id", stone_ids) \
            .execute().data or []
        for r in media_rows:
            if r["file_type"] in ("photo", "video", "animation"):
                eid = r["entity_id"]
                media_counts[eid] = media_counts.get(eid, 0) + 1
    except Exception:
        pass

    kb = InlineKeyboardBuilder()
    for s in data.data:
        count = media_counts.get(s["id"], 0)
        badge = f"📷{count}" if count else "📷—"
        kb.button(text=f"{badge} · {fmt_stone_btn(s)}", callback_data=f"view_stone_{s['id']}")
    kb.button(text="◀️ Назад", callback_data="action_view")
    kb.adjust(1)

    titles = {"natural": "🌿 Природные", "synthetic": "⚗️ Синтетика", "all": "📋 Все"}
    await callback.message.edit_text(f"👁 *{titles[filter_type]}* — выбери камень:",
                                     reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(F.data.startswith("view_filter_"))
async def view_filter_applied(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    filter_type = callback.data.replace("view_filter_", "")
    await _view_show_list(callback, filter_type)


@dp.callback_query(F.data.startswith("view_stone_"))
async def view_step2_card(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    stone_id = callback.data.replace("view_stone_", "")
    res = supabase.table("stones").select(
        "id,stone_code,stone_type,shape,carat,color,clarity,status,"
        "purchase_price,purchase_currency,purchase_price_usd,created_at,created_by,supplier_id"
    ).eq("id", stone_id).execute()

    if not res.data:
        await callback.message.edit_text("❌ Камень не найден.")
        return

    # Удаляем список камней, чтобы медиа шло первым чистым блоком
    try:
        await callback.message.delete()
    except Exception:
        pass

    await _send_stone_card(callback.message, res.data[0])


@dp.message(Command("stone"))
async def cmd_stone(message: Message, state: FSMContext):
    await state.clear()
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer("Использование: /stone <код>\n\nПример: /stone ST-2026-001")
        return

    code = args[1].strip().upper()
    res = supabase.table("stones").select(
        "id,stone_code,stone_type,shape,carat,color,clarity,status,"
        "purchase_price,purchase_currency,purchase_price_usd,created_at,created_by,supplier_id"
    ).ilike("stone_code", code).execute()

    if not res.data:
        await message.answer(f"❌ Камень `{code}` не найден.", parse_mode="Markdown")
        return

    await _send_stone_card(message, res.data[0])


# ============================================================
# Запуск
# ============================================================

# ============================================================
# Свободный текст → AI
# ============================================================

@dp.message(F.text & ~F.text.startswith("/"))
async def handle_ai(message: Message, state: FSMContext):
    if await state.get_state() is not None:
        return
    if not ANTHROPIC_API_KEY:
        await message.answer("⚠️ AI не настроен: добавь ANTHROPIC_API_KEY в переменные Railway.")
        return
    thinking = await message.answer("🤔")
    try:
        answer = await ask_claude(message.text)
        await thinking.edit_text(answer)
    except Exception as e:
        await thinking.edit_text(f"❌ Ошибка AI: {e}")


async def main():
    print("Бот запущен...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
